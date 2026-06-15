import csv
import io
import json
import os
import tempfile
from pathlib import Path
from flask import Flask, jsonify, request, send_from_directory, Response

from ledger.db import close_db, get_db, init_db, row_to_dict
from ledger.parser import parse_text
from ledger.reconciliation import compare_items
from ledger.services import create_order, delete_order, export_csv, list_orders, save_reconciliation, stats


def create_app(test_config=None):
    app = Flask(__name__, static_folder="static", static_url_path="")
    app.config.from_mapping(DATABASE=os.environ.get("LEDGER_DB", str(Path(app.root_path) / "data" / "ledger.db")))
    if test_config:
        app.config.update(test_config)
    init_db(app)
    app.teardown_appcontext(close_db)

    @app.route("/")
    def index():
        return send_from_directory(app.static_folder, "index.html")

    @app.post("/api/parse")
    def api_parse():
        payload = request.get_json(silent=True) or {}
        raw_text = payload.get("text") or ""
        return jsonify(parse_text(raw_text))

    @app.post("/api/orders")
    def api_create_order():
        payload = request.get_json(silent=True) or {}
        items = payload.get("items") or []
        if not items:
            return jsonify({"error": "items required"}), 400
        order = create_order(payload.get("source") or "manual", payload.get("raw_text") or "", items, payload.get("status") or "confirmed", bool(payload.get("final_override")))
        return jsonify(order), 201

    @app.get("/api/orders")
    def api_orders():
        return jsonify({"orders": list_orders(request.args.get("date"))})

    @app.delete("/api/orders/<int:order_id>")
    def api_delete_order(order_id):
        delete_order(order_id)
        return jsonify({"ok": True})

    @app.get("/api/stats")
    def api_stats():
        return jsonify(stats(request.args.get("range") or "day"))

    @app.get("/api/price-list")
    def api_price_list():
        rows = get_db().execute("SELECT * FROM price_list ORDER BY school_level, id").fetchall()
        return jsonify({"items": [row_to_dict(r) for r in rows]})

    @app.put("/api/price-list/<int:price_id>")
    def api_update_price(price_id):
        payload = request.get_json(silent=True) or {}
        db = get_db()
        db.execute("UPDATE price_list SET sell_price = ?, cost_price = ? WHERE id = ?", (float(payload["sell_price"]), float(payload["cost_price"]), price_id))
        db.commit()
        row = db.execute("SELECT * FROM price_list WHERE id = ?", (price_id,)).fetchone()
        return jsonify(row_to_dict(row))

    @app.post("/api/reconcile")
    def api_reconcile():
        payload = request.get_json(silent=True) or {}
        our_items = payload.get("our_items")
        if our_items is None and payload.get("our_text"):
            our_items = parse_text(payload.get("our_text"))["items"]
        if our_items is None and payload.get("our_date"):
            orders = list_orders(payload.get("our_date"), limit=100000)
            our_items = [item for order in orders for item in order["items"]]
        their_items = payload.get("their_items")
        if their_items is None:
            their_items = parse_text(payload.get("their_text") or "")["items"]
        result = compare_items(our_items or [], their_items or [])
        if payload.get("save"):
            result["id"] = save_reconciliation(payload.get("name") or "未命名对账", our_items or [], their_items or [], "matched" if result["summary"]["mismatched"] == 0 else "mismatch")
        return jsonify(result)

    @app.get("/api/reconciliations")
    def api_reconciliations():
        rows = get_db().execute("SELECT * FROM reconciliation_sets ORDER BY created_at DESC, id DESC").fetchall()
        return jsonify({"items": [row_to_dict(r) for r in rows]})

    @app.post("/api/export")
    def api_export():
        payload = request.get_json(silent=True) or {}
        content = export_csv(payload.get("date"))
        return Response(content, mimetype="text/csv; charset=utf-8", headers={"Content-Disposition": "attachment; filename=ledger.csv"})

    @app.post("/api/ocr")
    def api_ocr():
        uploaded = request.files.get("file")
        if not uploaded:
            return jsonify({"error": "file required"}), 400
        suffix = Path(uploaded.filename or "upload").suffix.lower()
        text = ""
        if suffix in {".txt", ".csv"}:
            raw = uploaded.read().decode("utf-8-sig", errors="ignore")
            if suffix == ".csv":
                rows = csv.reader(io.StringIO(raw))
                text = "\n".join(" ".join(cell for cell in row if cell) for row in rows)
            else:
                text = raw
        elif suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    uploaded.save(tmp.name)
                    path = tmp.name
                wb = load_workbook(path, data_only=True)
                lines = []
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        values = [str(v) for v in row if v is not None]
                        if values:
                            lines.append(" ".join(values))
                text = "\n".join(lines)
                os.unlink(path)
            except Exception as exc:
                return jsonify({"error": f"xlsx parse failed: {exc}"}), 400
        else:
            try:
                import pytesseract
                from PIL import Image
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix or ".png") as tmp:
                    uploaded.save(tmp.name)
                    path = tmp.name
                text = pytesseract.image_to_string(Image.open(path), lang="chi_sim+eng")
                os.unlink(path)
            except Exception as exc:
                return jsonify({"error": f"ocr failed: {exc}"}), 400
        parsed = parse_text(text)
        parsed["text"] = text
        return jsonify(parsed)

    return app


app = create_app()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
